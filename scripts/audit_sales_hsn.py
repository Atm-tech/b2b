from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


INPUT = Path(r"D:\AAPOORTI\Managment system\Sales managment\SALE 1-4-26 TO 26-7-26.xlsx")
OUTPUT = INPUT.with_name("SALE 1-4-26 TO 26-7-26 - HSN GST ERP FINAL.xlsx")

CBIC_RATE_SOURCE = "CBIC Notification 9/2025-Integrated Tax (Rate), effective 22-09-2025"
CBIC_EXEMPT_SOURCE = "CBIC Notification 10/2025-Integrated Tax (Rate), effective 22-09-2025"
CBIC_BEVERAGE_SOURCE = "CBIC Notification 1/2026-Integrated Tax (Rate), effective 01-05-2026"
DGFT_SOURCE = "DGFT ITC (HS), 2022 Schedule 1; six-digit HS subheading convention"
CBIC_SERVICE_SOURCE = "CBIC Scheme of Classification of Services; SAC 996813 local delivery services"


@dataclass(frozen=True)
class Classification:
    hsn6: str
    rate: int
    reason: str
    confidence: str = "HIGH"
    source: str = CBIC_RATE_SOURCE


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm(value: Any) -> str:
    return re.sub(r"[^A-Z0-9.+%&/-]+", " ", clean_text(value).upper()).strip()


def contains(text: str, *terms: str) -> bool:
    return any(term in text for term in terms)


def c(hsn6: str, rate: int, reason: str, confidence: str = "HIGH", source: str = CBIC_RATE_SOURCE) -> Classification:
    assert re.fullmatch(r"\d{6}", hsn6), hsn6
    return Classification(hsn6, rate, reason, confidence, source)


def pack_is_loose_or_over_25kg(text: str) -> bool:
    return "LOOSE" in text or bool(re.search(r"\b(?:30|40|50)\s*KG\b", text))


def classify(vertical: str, division: str, department: str, name: str, mrp: float | None) -> Classification | None:
    v, d, dept, n = map(norm, (vertical, division, department, name))
    text = f"{v} {d} {dept} {n}"

    # Exact product-master corrections for descriptions that do not expose their
    # underlying component/material. These are deliberately resolved as best
    # matches so the ERP import is complete.
    if n == "DELIVERY CHARGE":
        return c("996813", 18, "Local delivery service; 996813 is a six-digit SAC (service code), not a goods HSN.", "MEDIUM", CBIC_SERVICE_SOURCE)
    if n == "ACCESSORIES-PARTY ITEMS-MASKS":
        return c("950590", 18, "Party/festive mask classified as other festive, carnival or entertainment article.")
    if n.startswith("GROCERY COMBOPACK"):
        return c("210690", 5, "Mixed retail grocery combo classified as other food preparation for a single ERP product-master code.", "MEDIUM")
    if n == "GIFT PK BOMBAY DYEING":
        return c("630492", 5 if (mrp or 0) <= 2500 else 18, "Bombay Dyeing textile gift pack treated as other cotton furnishing article; rate follows the ₹2,500 textile sale-value threshold.", "MEDIUM")

    # Exempt and special-rate goods.
    if contains(n, "SANITARY PAD", "STAYFREE", "WHISPER", "SANITARY NAPKIN"):
        return c("961900", 0, "Sanitary towels/pads are exempt.", source=CBIC_EXEMPT_SOURCE)
    if contains(n, "COCA COLA", "COCA-COLA", "FANTA", "MIRINDA", "LIMCA", "SPRITE", "THUMS UP", "PEPSI", "7UP", "MOUNTAIN DEW"):
        return c("220210", 40, "Aerated/flavoured beverage with added sugar or sweetener.")
    if contains(n, "APPY FIZZ"):
        return c("220210", 40, "Carbonated fruit beverage.", source=f"{CBIC_RATE_SOURCE}; {CBIC_BEVERAGE_SOURCE}")
    if contains(n, "ENERGY DRINK", "STING ", "RED BULL"):
        return c("220299", 40, "Caffeinated/energy beverage.", source=f"{CBIC_RATE_SOURCE}; {CBIC_BEVERAGE_SOURCE}")
    if contains(n, "AUX CABLE", "AUX CUBLE", "USB CABLE"):
        return c("854442", 18, "Insulated electric conductor/cable fitted with connectors.")
    if contains(n, "CHARGER", "ACCESSORIES-ACCESSORIES-CHARGE"):
        return c("850440", 18, "Static converter/charger.")

    # Staples whose rate depends on retail packaging.
    loose = pack_is_loose_or_over_25kg(n)
    if dept in {"BASMATI", "RAW RICE", "SELA", "DUBRAJ", "RICE"} or re.search(r"\bRICE\b", n):
        return c("100630", 0 if loose else 5, "Rice; 5% only when pre-packaged and labelled up to 25 kg.", source=f"{CBIC_RATE_SOURCE}; {CBIC_EXEMPT_SOURCE}")
    if dept == "PULSES" or contains(n, " TOOR DAL", " TOOR DAAL", "ARHAR", "MOONG DAL", "URAD DAL", "MASOOR", "CHANA DAL", "KABULI CHANA", "RAJMA"):
        pulse_hsn = "071360" if contains(n, "TOOR", "ARHAR") else "071331" if contains(n, "MOONG", "URAD") else "071340" if "MASOOR" in n else "071320" if contains(n, "CHANA", "CHICKPEA") else "071333" if "RAJMA" in n else "071390"
        return c(pulse_hsn, 0 if loose else 5, "Dried pulse; 5% only when pre-packaged and labelled up to 25 kg.", source=f"{CBIC_RATE_SOURCE}; {CBIC_EXEMPT_SOURCE}")
    if dept in {"FLOUR", "ATTA"} or contains(n, " ATTA ", "FLOUR", "BESAN"):
        hsn = "110610" if "BESAN" in n else "110100" if contains(n, "ATTA", "WHEAT FLOUR", "MAIDA") else "110290"
        return c(hsn, 0 if loose else 5, "Flour; 5% only when pre-packaged and labelled up to 25 kg.", source=f"{CBIC_RATE_SOURCE}; {CBIC_EXEMPT_SOURCE}")
    if dept == "POHA" or " POHA" in n:
        return c("190490", 0 if loose else 5, "Flattened rice; 5% only when pre-packaged and labelled up to 25 kg.", source=f"{CBIC_RATE_SOURCE}; {CBIC_EXEMPT_SOURCE}")
    if dept == "SAGO" or contains(n, "SABUDANA", "SAGO"):
        return c("190300", 5, "Tapioca/sago preparation.")
    if dept == "CEREALS":
        cereal_rules = [
            (("BARBATI",), "071335"), (("DESI CHANA",), "071320"),
            (("JOWAR", "SORGHUM"), "100790"), (("PHALI DANA", "GROUNDNUT"), "120242"),
        ]
        for terms, hsn in cereal_rules:
            if contains(n, *terms):
                return c(hsn, 0 if loose else 5, "Raw cereal/pulse/oil seed; 5% only where the applicable entry is pre-packaged and labelled.", "MEDIUM", f"{CBIC_RATE_SOURCE}; {CBIC_EXEMPT_SOURCE}")
        return c("100890", 0 if loose else 5, "Other cereal; verify the exact grain.", "LOW", f"{CBIC_RATE_SOURCE}; {CBIC_EXEMPT_SOURCE}")
    if dept == "SUGAR" or "SUGAR" in n:
        return c("170199", 5, "Cane or beet sugar.")
    if dept == "SALT" or re.search(r"\bSALT\b", n):
        return c("250100", 0, "Common salt is exempt.", source=CBIC_EXEMPT_SOURCE)

    # Spices, seeds and dry fruits.
    if dept in {"SPICES", "WHOLE SPICES", "POWDER SPICES", "MASALA"} or "MASALA" in n:
        spice_rules = [
            (("TURMERIC", "HALDI"), "091030"), (("CHILLI", "CHILI", "MIRCH"), "090422"),
            (("CORIANDER", "DHANIA"), "090922"), (("CUMIN", "JEERA"), "090932"),
            (("FENNEL", "SOUNF"), "090962"), (("PEPPER", "KALI MIRCH"), "090412"),
            (("CARDAMOM", "ELAICHI"), "090832"), (("CLOVE", "LOUNG", "LAUNG"), "090720"),
            (("CINNAMON", "DALCHINI"), "090619"), (("NUTMEG", "JAIPHAL"), "090811"),
            (("MACE", "JAVITRI"), "090821"), (("AJWAIN",), "091099"),
            (("METHI", "FENUGREEK"), "091099"), (("MUSTARD", "RAI "), "120750"),
        ]
        for terms, hsn in spice_rules:
            if contains(n, *terms):
                return c(hsn, 5, "Single spice, ground/crushed unless the name states otherwise.")
        return c("091091", 5, "Mixed spice/masala.", "MEDIUM")
    if dept == "TIL" or re.search(r"\bTIL\b|\bSESAME\b", n):
        return c("120740", 5, "Sesame seeds other than seed quality.")
    if dept == "DRY FRUITS":
        dry_rules = [
            (("BADAM", "ALMOND"), "080212"), (("CASHEW", "KAJU"), "080132"),
            (("RAISIN", "KISHMISH"), "080620"), (("WALNUT", "AKHROT"), "080232"),
            (("PISTA", "PISTACHIO"), "080252"), (("DATE", "KHAJUR"), "080410"),
            (("FIG", "ANJEER"), "080420"),
        ]
        for terms, hsn in dry_rules:
            if contains(n, *terms):
                return c(hsn, 5, "Dried nut or fruit.")
        return c("081340", 5, "Other dried fruit/nut; verify exact botanical product.", "MEDIUM")

    # Edible oils and dairy.
    if dept in {"SOYA", "MUSTURED", "SUNFLOWER", "GROUNDNUT", "RICE BRAN", "COOKING MEDIUM", "VANASPATI", "MIX"} or " OIL" in n:
        if contains(n, "HAIR OIL", "BABY OIL", "BODY OIL", "MASSAGE OIL"):
            pass
        else:
            hsn = "150790" if contains(n, "SOYA", "SOYABEAN") else "151499" if contains(n, "MUSTARD", "SARSO") else "151219" if "SUNFLOWER" in n else "150890" if contains(n, "GROUNDNUT", "G.NUT") else "151590" if "RICE BRAN" in n else "151620" if "VANASPATI" in n else "151590"
            return c(hsn, 5, "Edible vegetable oil/fat.")
    if dept == "PURE GHEE" or "GHEE" in n:
        return c("040590", 5, "Ghee and other fats derived from milk.")
    if dept == "BUTTER" or re.search(r"\bBUTTER\b", n):
        return c("040510", 5, "Butter.")
    if dept == "BUTTER & CHEES" or "CHEESE" in n:
        return c("040690", 5, "Cheese.")
    if contains(n, "MILK ", " MILK") and not contains(n, "MILK POWDER", "MILK SHAKE", "MILK DRINK", "CHOCOLATE"):
        return c("040120", 0, "Fresh/UHT milk without added sugar is exempt.", source=CBIC_EXEMPT_SOURCE)
    if dept == "MILK POWDER":
        return c("040221", 5, "Milk powder.")
    if contains(n, "LASSI", "CHACH", "BUTTERMILK", "CURD", "YOGURT", "YOGHURT"):
        return c("040390", 5, "Pre-packaged fermented milk, curd, lassi or buttermilk.")
    if contains(n, "DAHI"):
        return c("040390", 5, "Pre-packaged curd/dahi.")
    if contains(n, "MILKMAID", "CONDENSED MILK"):
        return c("040299", 5, "Condensed milk.")

    # Prepared foods and drinks.
    if dept in {"BISCUITS", "MARIE", "BOURBON"} or contains(n, "BISCUIT", "COOKIE", "PARLE-G"):
        return c("190531", 5, "Sweet biscuits/cookies.")
    if dept in {"CAKE", "TOAST", "BAKERY"} and "FMCG-FOOD" in d:
        return c("190590", 5, "Cake, toast or other bakers' ware.")
    if dept == "WAFFERS":
        return c("190532", 5, "Waffles and wafers.")
    if dept in {"NAMKEEN", "READY TO EAT", "OTHER FRY ITEM", "FAST FOOD"}:
        return c("210690", 5, "Namkeen or other ready-to-eat food preparation.")
    if dept in {"READY TO COOK", "READY FOODS", "BASIC"}:
        return c("210690", 5, "Ready-to-cook/other food preparation.", "MEDIUM")
    if dept == "NOODLES" or contains(n, "NOODLE", "PASTA", "VERMICELLI"):
        return c("190230", 5, "Pasta/noodles.")
    if dept == "PAPAD" or "PAPAD" in n:
        return c("190590", 5, "Papad/other bakers' ware.")
    if dept == "CORNFLAKES" or contains(n, "CORN FLAKE", "CORNFLAKE"):
        return c("190410", 5, "Prepared cereal flakes.")
    if dept == "OATS":
        return c("110412", 5, "Rolled/flaked oats.")
    if dept in {"KETCHUP", "SAUCES"} or contains(n, "KETCHUP", "SAUCE"):
        return c("210320" if "TOMATO" in n or "KETCHUP" in n else "210390", 5, "Sauce, ketchup or condiment.")
    if dept == "PICKLES" or "PICKLE" in n:
        return c("200190", 5, "Vegetable/fruit preparation preserved by vinegar.")
    if dept == "VINEGAR" or "VINEGAR" in n:
        return c("220900", 18, "Vinegar and substitutes for vinegar.")
    if dept == "JAM" or contains(n, "JAM ", "MARMALADE"):
        return c("200799", 5, "Jam, fruit jelly or fruit preparation.")
    if dept in {"CONFECTIONARY", "SWEETS", "MOUTH FRESHNER"}:
        if contains(n, "CHOCOLATE", "KITKAT", "DAIRY MILK", "5 STAR", "PERK"):
            return c("180690", 5, "Chocolate preparation.")
        return c("170490", 5, "Sugar confectionery.")
    if dept == "CHOCOLATE":
        return c("180690", 5, "Chocolate or cocoa preparation.")
    if dept in {"HEALTH DRINKS", "HEALTH FOOD", "HEALTHY", "CUSTURED"}:
        return c("190190", 5, "Malt/cereal/milk-based food preparation.", "MEDIUM")
    if dept == "BABY FOOD" or contains(n, "POWERVITA"):
        return c("190110", 5, "Food preparation suitable for infants/young children or malt-based nutrition preparation.", "MEDIUM")
    if dept == "SOUP":
        return c("210410", 5, "Soup or broth preparation.")
    if dept == "COFFEE":
        return c("210111" if contains(n, "INSTANT", "NESCAFE", "BRU") else "090121", 5, "Coffee or coffee preparation.", "MEDIUM")
    if dept in {"TEA", "GREEN TEA", "TEA BAGS"}:
        return c("090230", 5, "Tea in retail packs not exceeding 3 kg.")
    if dept in {"WATER", "DRINKING WATER"}:
        return c("220110", 5, "Packaged drinking/mineral water.")
    if contains(n, "COCONUT WATER"):
        return c("200989", 5, "Pre-packaged tender coconut water.")
    if contains(n, "MAAZA", "FRUIT DRINK", "JUICE", "PAPER BOAT", "LITCHI DRINK"):
        return c("220299", 5, "Non-carbonated fruit-pulp/fruit-juice-based drink.", source=f"{CBIC_RATE_SOURCE}; {CBIC_BEVERAGE_SOURCE}")
    if contains(n, "SLICE ", "ICED LATTE", "ICED FRAPPE"):
        return c("220299", 5, "Non-carbonated fruit-pulp or milk-based beverage.", "MEDIUM", f"{CBIC_RATE_SOURCE}; {CBIC_BEVERAGE_SOURCE}")
    if contains(n, "ICED TEA", "NESTEA"):
        return c("210120", 5, "Tea-based preparation.")
    if dept in {"DRINKS & BREVRAGE", "BEVERAGE"}:
        return c("220299", 18, "Beverage description is insufficient to distinguish the 5% and 40% entries.", "LOW")
    if dept == "DOG FOOD":
        return c("230910", 18, "Dog or cat food put up for retail sale.", "MEDIUM")
    if dept == "FROZEN":
        return c("210690", 5, "Frozen prepared food.", "MEDIUM")
    if dept in {"FRUITS", "VEGETABLES"} or v == "F&V":
        return c("070999" if dept == "VEGETABLES" else "081090", 0, "Fresh fruit/vegetable; exact botanical subheading needs product detail.", "LOW", CBIC_EXEMPT_SOURCE)

    # Personal care and household chemicals.
    if dept == "SHAMPOO" or "SHAMPOO" in n:
        return c("330510", 5, "Shampoo.")
    if dept == "PASTE" or contains(n, "TOOTHPASTE", "TOOTH PASTE", "DANT KANTI"):
        return c("330610", 5, "Toothpaste.")
    if dept == "BRUSH" and contains(text, "FMCG-NON FOOD", "ORAL CARE"):
        return c("960321", 5, "Toothbrush.")
    if dept == "TALC" or contains(n, "TALC", "DERMI COOL", "NYCIL", "FACE POWDER"):
        return c("330491", 5, "Talcum/face powder.")
    if dept == "HAIR OIL" or "HAIR OIL" in n:
        return c("330590", 5, "Hair oil.")
    if dept in {"CONDITIONER", "HAIR CARE", "HAIR COLOR", "COLOR", "HAIR DIE", "HAIR GEL", "HAIR SERUM", "HAIR CREAM"}:
        return c("330590", 18, "Hair preparation other than shampoo or hair oil.")
    if dept in {"FACE CREAM", "FACE CARE", "CREAM", "LOTION", "MOISTURIZER", "BODY CARE", "FACE PACK", "SKIN TONNER", "EYE CARE", "LIP CARE", "BLEACH"}:
        return c("330499", 18, "Skin-care/cosmetic preparation other than talcum or face powder.")
    if dept == "FACE WASH" or "FACE WASH" in n:
        return c("340130", 18, "Liquid/cream skin-washing preparation.")
    if dept in {"BODY WASH", "HAND WASH"} or contains(n, "BODY WASH", "HANDWASH", "HAND WASH", "SHOWER GEL"):
        return c("340130", 18, "Liquid skin-washing preparation.")
    if contains(n, "BATH SOAP", "TOILET SOAP", "LIRIL SOAP", "DETTOL SOAP", "LUX SOAP", "SANTOOR SOAP") and not contains(n, "DETERGENT"):
        return c("340111", 5, "Toilet soap bar.")
    if dept in {"CAKE", "BAR"} and "FMCG-NON FOOD" in d:
        return c("340119", 18, "Detergent/cleaning bar; verify if the item is instead toilet soap.", "MEDIUM")
    if dept in {"HOME CARE", "CLEANER", "FLOOR CLEANER", "TOILET CLEANER", "KITCHEN CARE", "CLOTH CLEANER", "LIQUID"}:
        if contains(n, "MOSQUITO", "COCKROACH", "INSECT", "ALL OUT", "HIT "):
            return c("380891", 18, "Retail insecticide/repellent.")
        return c("340250", 18, "Retail cleaning/washing preparation.", "MEDIUM")
    if dept in {"PESTICIDES", "MOSQUITO"} or contains(n, "MOSQUITO", "COCKROACH", "INSECT KILLER", "ALL OUT", "HIT "):
        return c("380891", 18, "Retail insecticide/repellent.")
    if dept in {"INCENT STICK", "DHOOP BATTI"} or contains(n, "AGARBATTI", "INCENSE", "DHOOP"):
        return c("330741", 5, "Agarbatti/dhoop or other odoriferous preparation operating by burning.")
    if dept in {"AIR FRESHNER", "ROOM FRESHNER"}:
        return c("330749", 18, "Prepared room deodoriser/air freshener.")
    if dept in {"DEO", "ROLL-ON"}:
        return c("330720", 18, "Personal deodorant.")
    if dept in {"PERFUMES", "PERFUME"}:
        return c("330300", 18, "Perfume/toilet water.")
    if dept in {"SHAVING CREAM", "AFTER SHAVE"}:
        return c("330710", 5, "Shaving cream/lotion or after-shave lotion.")
    if dept == "RAZOR":
        return c("821210", 18, "Razor.")
    if contains(n, "GILLETTE", " RZR"):
        return c("821210", 18, "Razor.")
    if dept == "BLADE":
        return c("821220", 18, "Razor blade.")
    if dept == "BATTRIES" or contains(n, "BATTERY", "BATTERIES", "DURACELL"):
        return c("850610", 18, "Primary cells/batteries.")
    if dept == "POLISH":
        return c("340510", 18, "Footwear/furniture polish.")
    if dept in {"SANITORY", "BABY CARE", "BABY PRODUCTS", "BABY WIPES"}:
        if contains(n, "DIAPER", "NAPPIE", "BABY NAPKIN"):
            return c("961900", 5, "Baby napkin/diaper.")
        if "WIPES" in n:
            return c("340119", 18, "Impregnated cleansing wipes.", "MEDIUM")
        return c("330499", 18, "Baby/personal-care preparation; exact composition should be checked.", "LOW")
    if dept in {"TISSUE PAPER", "NAPKIN PAPER"}:
        return c("481820", 18, "Paper tissues/handkerchiefs.")
    if dept == "MEHNDI":
        return c("140490", 5, "Mehendi paste/powder.", "MEDIUM")
    if dept == "CAMPHOR":
        return c("291429", 18, "Camphor.")
    if dept == "MATCH BOX":
        return c("360500", 5, "Matches.")
    if contains(n, "ARIEL", "DETERGENT POWDER", "WASHING POWDER"):
        return c("340250", 18, "Retail detergent/washing preparation.")
    if contains(n, "NYCIL", "TELC ", "TALC "):
        return c("330491", 5, "Talcum powder.")
    if dept == "HAIR REMOVER" or contains(n, "H/REMOVER", "HAIR REMOVER"):
        return c("330790", 18, "Depilatory preparation.")
    if contains(n, "PEPSODENT"):
        return c("330610", 5, "Toothpaste.")
    if dept == "DUSTER" or contains(n, "DUSTER -"):
        return c("630710", 5, "Floor/dusting cloth of textile material, sale value not exceeding ₹2,500.", "MEDIUM")
    if contains(n, "PAPER CUP"):
        return c("482369", 18, "Paper/paperboard cup or other tableware.")
    if dept == "UNDERGARMENTS":
        return c("610711", 5 if (mrp or 0) <= 2500 else 18, "Knitted undergarment; rate based on ₹2,500 sale-value threshold.", "MEDIUM")
    if dept == "CARRY BAG":
        return c("392329", 18, "Plastic carry bag; verify material.", "LOW")

    # Stationery and accessories.
    if contains(dept, "STATIONARY", "STATIONERY", "BOOKS") or contains(n, "STATIONARY", "STATIONERY", "ACCESSORIES-BOOKS-") or d in {"STATIONARY", "STATIONERY"} or dept in {"NOTE BOOK", "PENCIL", "PEN", "ERASER", "SHARPNER", "WAX COLOUR", "CLAY", "DIARY", "GEOMETRY BOX", "MARKER", "PEN STAND", "PENCIL POUCH", "SCALE"}:
        rules = [
            (("NOTE BOOK", "NOTEBOOK", "EXERCISE BOOK"), "482020", 0, "Notebook/exercise book."),
            (("PENCIL POUCH", "POUCH"), "420292", 18, "Pencil/stationery pouch; verify outer material."),
            (("PEN STAND",), "392610", 18, "Plastic office/stationery article; verify material."),
            (("PENCIL",), "960910", 0, "Pencil."),
            (("CRAYON", "WAX COLOUR", "PASTEL"), "960990", 0, "Crayon/pastel."),
            (("ERASER",), "401692", 5, "Rubber eraser."),
            (("SHARPNER",), "821410", 5, "Pencil sharpener."),
            (("MARKER", "SKETCH PEN"), "960820", 18, "Felt-tipped/porous-tipped pen or marker."),
            (("DIARY", "REGISTER"), "482010", 18, "Register/diary."),
            (("FILE", "FOLDER"), "482030", 18, "Binder/file cover."),
            (("COVER",), "482030", 18, "Paperboard/plastic file or document cover; verify material."),
            (("CLAY",), "340700", 18, "Modelling paste/clay."),
            (("GEOMETRY BOX", "COMPASS"), "731029", 5, "Mathematical/geometry box."),
            (("SCALE", "RULER"), "901780", 18, "Hand-held measuring instrument/ruler."),
            (("PAPER",), "480256", 18, "Writing/printing paper; verify grammage and size."),
            (("PEN",), "960810", 18, "Ball-point pen."),
        ]
        for terms, hsn, rate, reason in rules:
            if contains(n, *terms) or contains(dept, *terms):
                return c(hsn, rate, reason, "MEDIUM", CBIC_EXEMPT_SOURCE if rate == 0 else CBIC_RATE_SOURCE)
        return c("482090", 18, "Other paper stationery; verify material.", "LOW")

    # Apparel and footwear. Six-digit code is provisional where fibre/material is absent.
    if d in {"MENS", "LADIES", "BOYS", "GIRLS", "INFANT"} or v == "HYPER" and dept in {
        "SHIRTS", "T-SHIRTS", "T-SHIRT", "JEANS", "TROUSERS", "CASUAL TROUSERS", "TOP", "KURTI",
        "PYJAMA", "PAJAMA", "LOWER", "LEGGING", "FROCK", "BRIEF", "VEST", "CAPRI", "SAREE",
        "SALWAR SUIT", "NIGHT SUIT", "JACKET", "BLAZER", "KURTA", "BOXER"
    }:
        apparel_rules = {
            "T-SHIRTS": "610910", "T-SHIRT": "610910", "SHIRTS": "620520", "JEANS": "620342",
            "TROUSERS": "620342", "CASUAL TROUSERS": "620342", "TOP": "620630", "KURTI": "621142",
            "PYJAMA": "620721", "PAJAMA": "620721", "LOWER": "610343", "LEGGING": "610463",
            "FROCK": "620442", "BRIEF": "610711", "VEST": "610910", "CAPRI": "620342",
            "SAREE": "520849", "SALWAR SUIT": "621142", "NIGHT SUIT": "620721", "JACKET": "620332",
            "BLAZER": "620332", "KURTA": "621132", "BOXER": "610711",
        }
        hsn = apparel_rules.get(dept, "621149")
        rate = 5 if (mrp or 0) <= 2500 else 18
        return c(hsn, rate, "Apparel rate is based on sale value threshold ₹2,500; fibre/knit detail is inferred from the generic description.", "LOW")
    if d == "FOOTMART" or dept in {"CHAPPAL", "SHOES", "SANDLE", "CROCS", "BABY BOOT"}:
        hsn = "640220" if dept in {"CHAPPAL", "SANDLE"} else "640299" if dept == "CROCS" else "640419"
        rate = 5 if (mrp or 0) <= 2500 else 18
        return c(hsn, rate, "Footwear rate is based on sale value threshold ₹2,500; material must be checked for exact subheading.", "LOW")
    if contains(n, "HANKY", "HANDKERCHIEF"):
        return c("621320", 5 if (mrp or 0) <= 2500 else 18, "Cotton handkerchief; verify fibre.", "LOW")
    if dept in {"HAND BAG", "TRAVELLING BAG", "LAPTOP BAGS", "SCHOOL BAG", "BELT", "CAP", "UMBRELLA"} or contains(n, "HAND BAG", "TRAVELLING BAG", "LAPTOP BAG", "SCHOOL BAG", "PICNIC BAG", "SHOPPING BAG", "BLANKET BAG"):
        if "UMBRELLA" in dept or "UMBRELLA" in n:
            return c("660199", 5, "Umbrella.")
        if dept == "CAP":
            return c("650500", 5, "Textile cap/hat.", "MEDIUM")
        if dept == "BELT":
            return c("420330", 18, "Belt; assumes leather/composition leather.", "LOW")
        return c("420292", 18, "Bag with outer surface of textile/plastic sheeting; material must be checked.", "LOW")

    # Household goods.
    if d == "HOUSEHOLD" or "HOUSEHOLD" in text:
        if contains(n, "STEEL", "STAINLESS", "SS "):
            return c("732393", 5, "Stainless-steel table/kitchen/household article.")
        if contains(n, "ALUMINIUM", "ALU "):
            return c("761510", 5, "Aluminium table/kitchen/household article.")
        if contains(n, "GLASS", "CROCKERY", "DINNER SET", "CUP SET"):
            return c("701349", 18, "Table/kitchen glassware; verify ceramic versus glass.", "LOW")
        if contains(n, "PLASTIC", "BOTTLE", "CONTAINER", "LUNCH BOX", "BUCKET", "BASKET", "DUSTBIN", "MUG", "TUB", "SOAP CASE", "TRAY", "JAR"):
            return c("392490", 18, "Plastic household article.", "MEDIUM")
        if contains(n, "KNIFE"):
            return c("821192", 18, "Kitchen/table knife.")
        if contains(n, "SCISSOR"):
            return c("821300", 18, "Scissors.")
        if contains(n, "SPOON", "FORK", "CUTLERY"):
            return c("821599", 18, "Spoon/fork or similar cutlery; material should be checked.", "LOW")
        if contains(n, "LOCK"):
            return c("830110", 18, "Padlock/lock.")
        if contains(n, "BROOM"):
            return c("960310", 0, "Broom of twigs/vegetable material; verify construction.", "LOW", CBIC_EXEMPT_SOURCE)
        if contains(n, "BRUSH", "WIPER", "MOP"):
            return c("960390", 5, "Other broom/brush/mop.", "LOW")
        return c("392490", 18, "Generic household article; provisional plastic classification—verify material.", "LOW")

    # Electrical/electronic and toys.
    if v == "ELECTRONICS" or d == "ELECTRONICS" or d == "ELECTRONICS":
        electronic_rules = [
            (("BULB",), "853952"), (("MIXER", "BLENDER"), "850940"), (("IRON",), "851640"),
            (("KETTLE",), "851679"), (("COOKER",), "851660"), (("TOASTER",), "851672"),
        ]
        for terms, hsn in electronic_rules:
            if contains(n, *terms) or contains(dept, *terms):
                return c(hsn, 18, "Electrical household appliance.")
        return c("854370", 18, "Other electrical apparatus; exact function should be checked.", "LOW")
    if contains(d, "TOYS") or contains(dept, "TOYS", "SPORTS") or " TOY" in n:
        electronic = contains(n, "ELECTRONIC", "BATTERY", "REMOTE")
        return c("950300", 18 if electronic else 5, "Toy; electronic toys are 18%, other listed toys are 5%.", "LOW")

    return None


def normalize_original_hsn(value: Any) -> str:
    digits = re.sub(r"\D", "", clean_text(value))
    if len(digits) == 8:
        return digits[:6]
    if len(digits) == 7:
        return digits.zfill(8)[:6]
    if len(digits) == 6:
        return digits
    generic = {
        "910": "091099", "1905": "190590", "2105": "210500", "2106": "210690",
        "3304": "330499", "3307": "330790", "3401": "340119", "3402": "340250",
        "3924": "392490", "4202": "420292", "6103": "610399", "6203": "620399",
    }
    return generic.get(digits, "")


def number(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def rate_value(value: Any) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def format_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    widths = {
        "A": 16, "B": 18, "C": 20, "D": 42, "E": 18, "F": 16, "G": 14,
        "H": 14, "I": 14, "J": 20, "K": 22, "L": 20, "M": 55, "N": 55,
    }
    for letter, width in widths.items():
        if ws.max_column >= ord(letter) - 64:
            ws.column_dimensions[letter].width = width


def add_table(ws, name: str) -> None:
    if ws.max_row < 2:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)


def main() -> None:
    source_wb = load_workbook(INPUT, read_only=True, data_only=True)
    source_ws = source_wb.active
    headers = [clean_text(cell.value) for cell in source_ws[1]]
    index = {header: pos for pos, header in enumerate(headers)}
    required = {"VERTICAL", "DIVISION", "DEPARTMENT", "HSNCODE", "ARTICLE NAME", "MRP", "TAX%"}
    missing = sorted(required - set(index))
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")

    raw_rows: list[tuple[Any, ...]] = []
    groups: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        name = clean_text(row[index["ARTICLE NAME"]])
        if not name or norm(name) == "GRAND TOTAL":
            continue
        raw_rows.append(row)
        groups[norm(name)].append(row)

    audit: dict[str, dict[str, Any]] = {}
    for key, rows in groups.items():
        display_name = Counter(clean_text(r[index["ARTICLE NAME"]]) for r in rows).most_common(1)[0][0]
        vertical = Counter(clean_text(r[index["VERTICAL"]]) for r in rows).most_common(1)[0][0]
        division = Counter(clean_text(r[index["DIVISION"]]) for r in rows).most_common(1)[0][0]
        department = Counter(clean_text(r[index["DEPARTMENT"]]) for r in rows).most_common(1)[0][0]
        mrps = [number(r[index["MRP"]]) for r in rows if number(r[index["MRP"]]) is not None]
        representative_mrp = max(mrps) if mrps else None
        original_hsns = Counter(clean_text(r[index["HSNCODE"]]) or "(blank)" for r in rows)
        original_rates = Counter(rate_value(r[index["TAX%"]]) for r in rows)
        classification = classify(vertical, division, department, display_name, representative_mrp)

        normalized_hsns = {normalize_original_hsn(value) for value in original_hsns if value != "(blank)"}
        normalized_hsns.discard("")
        fallback = Counter(
            normalize_original_hsn(r[index["HSNCODE"]])
            for r in rows
            if normalize_original_hsn(r[index["HSNCODE"]])
        ).most_common(1)
        if classification is None:
            fallback_hsn = fallback[0][0] if fallback else ""
            fallback_rate = Counter(x for x in (rate_value(r[index["TAX%"]]) for r in rows) if x in {0, 5, 18, 40}).most_common(1)
            classification = Classification(
                fallback_hsn,
                fallback_rate[0][0] if fallback_rate else 18,
                "No sufficiently specific product rule; retained the most common valid source value provisionally.",
                "LOW",
                DGFT_SOURCE,
            )

        original_rate_set = {x for x in original_rates if x is not None}
        hsn_matches = bool(classification.hsn6) and normalized_hsns == {classification.hsn6}
        rate_matches = original_rate_set == {classification.rate}
        conflicted = len(original_hsns) > 1 or len(original_rate_set) > 1
        if hsn_matches and rate_matches and not conflicted:
            status = "OK"
        else:
            status = "WRONG - CORRECTED"

        audit[key] = {
            "vertical": vertical,
            "division": division,
            "department": department,
            "name": display_name,
            "row_count": len(rows),
            "original_hsn": "; ".join(f"{value} ({count})" for value, count in original_hsns.most_common()),
            "original_tax": "; ".join(f"{value if value is not None else '(blank)'}% ({count})" for value, count in original_rates.most_common()),
            "hsn6": classification.hsn6,
            "rate": classification.rate,
            "status": status,
            "confidence": "BEST MATCH" if classification.confidence == "LOW" else "RULE MATCH" if classification.confidence == "MEDIUM" else "EXACT RULE",
            "reason": classification.reason,
            "source": classification.source,
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "Product Audit"
    product_headers = [
        "VERTICAL", "DIVISION", "DEPARTMENT", "ARTICLE NAME", "SOURCE ROW COUNT",
        "ORIGINAL HSN VALUES", "ORIGINAL TAX VALUES", "CORRECT HSN 6 DIGIT",
        "CORRECT TAX%", "AUDIT STATUS", "CLASSIFICATION BASIS", "AUDIT REASON", "SOURCE",
    ]
    ws.append(product_headers)
    fills = {
        "OK": PatternFill("solid", fgColor="C6EFCE"),
        "WRONG - CORRECTED": PatternFill("solid", fgColor="FFC7CE"),
    }
    for row_number, item in enumerate(sorted(audit.values(), key=lambda x: (x["status"], x["division"], x["department"], x["name"])), start=2):
        ws.append([
            item["vertical"], item["division"], item["department"], item["name"], item["row_count"],
            item["original_hsn"], item["original_tax"], item["hsn6"], item["rate"],
            item["status"], item["confidence"], item["reason"], item["source"],
        ])
        ws.cell(row_number, 10).fill = fills[item["status"]]
        ws.cell(row_number, 10).font = Font(bold=True)
        ws.cell(row_number, 8).number_format = "@"
        for column in (4, 6, 7, 12, 13):
            ws.cell(row_number, column).alignment = Alignment(vertical="top", wrap_text=True)
    format_sheet(ws)
    ws.column_dimensions["D"].width = 44
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["L"].width = 62
    ws.column_dimensions["M"].width = 62
    add_table(ws, "ProductAudit")

    erp = wb.create_sheet("ERP Import")
    erp.append(["ARTICLE NAME", "HSN 6 DIGIT", "GST %"])
    erp_rows = list(audit.values())
    for item in sorted(erp_rows, key=lambda x: x["name"]):
        erp.append([item["name"], item["hsn6"], item["rate"]])
        erp.cell(erp.max_row, 2).number_format = "@"
    format_sheet(erp)
    erp.column_dimensions["A"].width = 55
    erp.column_dimensions["B"].width = 18
    erp.column_dimensions["C"].width = 12
    add_table(erp, "ErpImport")

    cleaned = wb.create_sheet("Cleaned Sales")
    cleaned_headers = headers[:]
    cleaned_headers[index["HSNCODE"]] = "HSNCODE 6 DIGIT"
    cleaned_headers[index["TAX%"]] = "CORRECT TAX%"
    cleaned_headers.extend(["ORIGINAL HSNCODE", "ORIGINAL TAX%", "AUDIT STATUS", "AUDIT NOTE"])
    cleaned.append(cleaned_headers)
    cleaned_status_column = len(cleaned_headers) - 1
    for row_number, row in enumerate(raw_rows, start=2):
        values = list(row)
        item = audit[norm(row[index["ARTICLE NAME"]])]
        original_hsn = values[index["HSNCODE"]]
        original_tax = values[index["TAX%"]]
        values[index["HSNCODE"]] = item["hsn6"]
        values[index["TAX%"]] = item["rate"]
        source_hsn = normalize_original_hsn(original_hsn)
        source_rate = rate_value(original_tax)
        row_correct = bool(item["hsn6"]) and source_hsn == item["hsn6"] and source_rate == item["rate"]
        row_status = "OK" if row_correct and item["status"] == "OK" else item["status"]
        values.extend([original_hsn, original_tax, row_status, item["reason"]])
        cleaned.append(values)
        cleaned.cell(row_number, cleaned_status_column).fill = fills[row_status]
        cleaned.cell(row_number, cleaned_status_column).font = Font(bold=True)
        cleaned.cell(row_number, index["HSNCODE"] + 1).number_format = "@"
    format_sheet(cleaned)
    for i in range(1, cleaned.max_column + 1):
        cleaned.column_dimensions[get_column_letter(i)].width = min(45, max(12, cleaned.column_dimensions[get_column_letter(i)].width or 12))
    cleaned.column_dimensions[get_column_letter(index["ARTICLE NAME"] + 1)].width = 44
    cleaned.column_dimensions[get_column_letter(cleaned.max_column)].width = 60
    add_table(cleaned, "CleanedSales")

    summary = wb.create_sheet("Audit Summary", 0)
    summary.append(["AUDIT ITEM", "RESULT"])
    status_counts = Counter(item["status"] for item in audit.values())
    summary_rows = [
        ("Input file", str(INPUT)),
        ("Output file", str(OUTPUT)),
        ("Original sales rows", len(raw_rows)),
        ("Exact duplicate transaction rows removed", 0),
        ("Distinct article values after normalization", len(audit)),
        ("Repeated product-master values consolidated", len(raw_rows) - len(audit)),
        ("Products already correct", status_counts["OK"]),
        ("Products marked WRONG - CORRECTED", status_counts["WRONG - CORRECTED"]),
        ("Products left ambiguous", 0),
        ("Products ready in ERP Import sheet", len(erp_rows)),
        ("Important", "Historical TAXAMT/TAXABLEAMT/NETAMT values are retained; this audit does not rewrite posted invoice amounts."),
        ("HSN rule", "Every corrected HSN/SAC is exactly six digits; every article is included in the ERP Import sheet."),
        ("GST rate rule", "Corrected rates use only the current applicable 0%, 5%, 18% and 40% structure. No corrected 12% value remains."),
        ("GST basis", CBIC_RATE_SOURCE),
        ("Exemption basis", CBIC_EXEMPT_SOURCE),
        ("Beverage amendment", CBIC_BEVERAGE_SOURCE),
        ("HS nomenclature", DGFT_SOURCE),
    ]
    for row in summary_rows:
        summary.append(row)
    format_sheet(summary)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 110
    for row in summary.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    sources = wb.create_sheet("Sources & Notes")
    sources.append(["SOURCE", "URL / NOTE"])
    sources.append([CBIC_RATE_SOURCE, "https://taxinformation.cbic.gov.in/view-pdf/1010431/ENG/Notifications"])
    sources.append([CBIC_EXEMPT_SOURCE, "GST exemption schedule; non-prepackaged staples and specified exempt goods."])
    sources.append([CBIC_BEVERAGE_SOURCE, "https://taxinformation.cbic.gov.in/"])
    sources.append([DGFT_SOURCE, "https://content.dgft.gov.in/Website/General_Notes_regarding_Import_Policy_20225_updated.pdf"])
    sources.append([CBIC_SERVICE_SOURCE, "https://cbic-gst.gov.in/hindi/pdf/central-tax-rate/Notification11-CGST-Annexure.pdf"])
    sources.append(["Best-match rule", "Where the article description omits material, composition, knitting method or exact intended use, a single six-digit best-match code was assigned from the available product name/category so the ERP list has no ambiguous or blank rows."])
    format_sheet(sources)
    sources.column_dimensions["A"].width = 68
    sources.column_dimensions["B"].width = 110

    invalid_hsn = [item["name"] for item in audit.values() if not re.fullmatch(r"\d{6}", item["hsn6"])]
    invalid_rate = [item["name"] for item in audit.values() if item["rate"] not in {0, 5, 18, 40}]
    if invalid_hsn or invalid_rate:
        raise RuntimeError(f"Validation failed: invalid HSN/SAC={len(invalid_hsn)}, invalid rate={len(invalid_rate)}")

    wb.save(OUTPUT)
    print(f"OUTPUT={OUTPUT}")
    print(f"SALES_ROWS={len(raw_rows)}")
    print(f"DISTINCT_PRODUCTS={len(audit)}")
    for status, count in sorted(status_counts.items()):
        print(f"{status}={count}")


if __name__ == "__main__":
    main()
