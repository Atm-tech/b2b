from __future__ import annotations

import hashlib
import json
import os
import re
from collections import defaultdict
from datetime import datetime, time, timedelta

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import RealDictCursor


ROOT = os.path.dirname(os.path.dirname(__file__))
ENV_PATH = os.path.join(ROOT, ".env")
WORKBOOK_PATH = r"C:\Users\Windows11\Downloads\SALE INVOICE 28-MAR TO 28-APR-26.xlsx"
BACKUP_PATH = os.path.join(ROOT, "data", f"db-backup-before-sales-import-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json")

WAREHOUSE_MAP = {
    "GP": "GOVINDPURA",
    "GOVINDPURA": "GOVINDPURA",
    "C21W": "C21",
    "C21": "C21",
}


def clean_text(value: object) -> str:
    text = "" if value is None else str(value)
    return " ".join(text.replace("\xa0", " ").split()).strip()


def as_float(value: object) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def load_database_url() -> str:
    with open(ENV_PATH, "r", encoding="utf-8") as handle:
        env_text = handle.read()
    return env_text.split("DATABASE_URL=")[1].splitlines()[0].strip()


def parse_date(value: object) -> datetime:
    if isinstance(value, datetime):
        date_obj = value
    else:
        text = clean_text(value)
        date_obj = None
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                date_obj = datetime.strptime(text, fmt)
                break
            except ValueError:
                pass
        if date_obj is None:
            raise ValueError(f"Unsupported date value: {value!r}")
    if date_obj.hour == 0 and date_obj.minute == 0 and date_obj.second == 0:
        return datetime.combine(date_obj.date(), time(12, 0, 0))
    return date_obj


def shop_id(name: str) -> str:
    normalized = clean_text(name).upper()
    digest = hashlib.md5(normalized.encode("utf-8")).hexdigest()[:8].upper()
    token = re.sub(r"[^A-Z0-9]+", "-", normalized).strip("-")[:40] or "SHOP"
    return f"SHP-{token}-{digest}"


def serialize_rows(cur: RealDictCursor, query: str, params: list[object] | None = None) -> list[dict[str, object]]:
    cur.execute(query, params or [])
    return [dict(row) for row in cur.fetchall()]


def to_jsonable(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def make_sales_line_id(invoice: str, index: int) -> str:
    token = re.sub(r"[^A-Za-z0-9]+", "-", invoice).strip("-")
    return f"IMP-SO-{token}-{index:03d}"


def make_opening_lot_id(warehouse_id: str, product_sku: str) -> str:
    digest = hashlib.md5(f"{warehouse_id}::{product_sku}".encode("utf-8")).hexdigest()[:8].upper()
    return f"OPEN-{warehouse_id}-{digest}"


def consume_inventory(cur: RealDictCursor, warehouse_id: str, product_sku: str, quantity: float) -> None:
    remaining = quantity
    cur.execute(
        """
        SELECT lot_id, quantity_available
        FROM inventory_lots
        WHERE warehouse_id = %s AND product_sku = %s AND quantity_available > 0
        ORDER BY created_at ASC, lot_id ASC
        """,
        [warehouse_id, product_sku],
    )
    lots = cur.fetchall()
    for lot in lots:
        if remaining <= 0:
            break
        available = as_float(lot["quantity_available"])
        move = min(available, remaining)
        cur.execute(
            "UPDATE inventory_lots SET quantity_available = %s WHERE lot_id = %s",
            [round(available - move, 6), lot["lot_id"]],
        )
        remaining = round(remaining - move, 6)
    if remaining > 0.000001:
        raise ValueError(f"Insufficient inventory for {product_sku} at {warehouse_id}. Short by {remaining}.")


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Sheet1"]

    rows: list[dict[str, object]] = []
    for row_no in range(2, sheet.max_row + 1):
        invoice = clean_text(sheet.cell(row_no, 3).value)
        if not invoice:
            continue
        warehouse_code = clean_text(sheet.cell(row_no, 5).value).upper()
        warehouse_id = WAREHOUSE_MAP.get(warehouse_code)
        if not warehouse_id:
            raise ValueError(f"Unknown warehouse code {warehouse_code!r} on row {row_no}.")
        rows.append(
            {
                "barcode": clean_text(sheet.cell(row_no, 1).value),
                "product_name": clean_text(sheet.cell(row_no, 2).value),
                "invoice": invoice,
                "shop_name": clean_text(sheet.cell(row_no, 4).value),
                "warehouse_id": warehouse_id,
                "date": parse_date(sheet.cell(row_no, 6).value),
                "qty": round(as_float(sheet.cell(row_no, 7).value), 6),
                "amount": round(as_float(sheet.cell(row_no, 8).value), 2),
                "bill_type": clean_text(sheet.cell(row_no, 9).value).upper() or "NON GST",
            }
        )

    database_url = load_database_url()
    conn = psycopg2.connect(database_url)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        backup = {
            "sales_orders": serialize_rows(cur, "SELECT * FROM sales_orders"),
            "sales_payments": serialize_rows(cur, "SELECT * FROM payments WHERE side = 'Sales'"),
            "sales_ledger": serialize_rows(cur, "SELECT * FROM ledger_entries WHERE side = 'Sales'"),
            "sales_tasks": serialize_rows(cur, "SELECT * FROM delivery_tasks WHERE side = 'Sales'"),
            "sales_dockets": serialize_rows(cur, "SELECT * FROM delivery_dockets"),
            "sales_consignments": serialize_rows(cur, "SELECT * FROM delivery_consignments"),
            "shops": serialize_rows(cur, "SELECT * FROM counterparties WHERE type = 'Shop'"),
            "products": serialize_rows(cur, "SELECT * FROM products"),
            "inventory_lots": serialize_rows(cur, "SELECT * FROM inventory_lots"),
        }
        with open(BACKUP_PATH, "w", encoding="utf-8") as handle:
            json.dump(backup, handle, default=to_jsonable, indent=2)

        cur.execute("SELECT id, username, full_name, role FROM users ORDER BY id")
        users = cur.fetchall()
        sales_user = next((row for row in users if row["role"] == "Sales"), None)
        if not sales_user:
            raise ValueError("No Sales user found.")

        cur.execute("DELETE FROM delivery_dockets")
        cur.execute("DELETE FROM delivery_consignments")
        cur.execute("DELETE FROM delivery_tasks WHERE side = 'Sales'")
        cur.execute("DELETE FROM payments WHERE side = 'Sales'")
        cur.execute("DELETE FROM ledger_entries WHERE side = 'Sales'")
        cur.execute("DELETE FROM sales_orders")
        cur.execute("DELETE FROM counterparties WHERE type = 'Shop'")

        cur.execute(
            """
            SELECT sku, name, default_gst_rate, default_tax_mode, allowed_warehouse_ids_json, barcode, category, rsp, mrp
            FROM products
            """
        )
        products = {str(row["sku"]): dict(row) for row in cur.fetchall()}

        product_usage: dict[str, set[str]] = defaultdict(set)
        product_barcodes: dict[str, str] = {}
        product_gst_seen: dict[str, bool] = defaultdict(bool)
        shops: set[str] = set()
        invoice_groups: dict[str, list[dict[str, object]]] = defaultdict(list)
        invoice_meta: dict[str, dict[str, object]] = {}

        for row in rows:
            product_name = str(row["product_name"])
            product_usage[product_name].add(str(row["warehouse_id"]))
            product_barcodes[product_name] = str(row["barcode"])
            if str(row["bill_type"]) == "GST":
                product_gst_seen[product_name] = True
            shops.add(str(row["shop_name"]))
            invoice_groups[str(row["invoice"])].append(row)
            invoice_meta.setdefault(
                str(row["invoice"]),
                {
                    "shop_name": str(row["shop_name"]),
                    "date": row["date"],
                },
            )

        for name in sorted(product_usage):
            if name not in products:
                gst_rate = 5 if product_gst_seen[name] else 0
                tax_mode = "Exclusive" if product_gst_seen[name] else "NA"
                cur.execute(
                    """
                    INSERT INTO products (
                      sku, name, division, department, section_name, category, unit, default_weight_kg, tolerance_kg,
                      tolerance_percent, allowed_warehouse_ids_json, slabs_json, remarks, category_6, site_name, barcode,
                      supplier_name, hsn_code, article_name, item_name, brand, short_name, size, rsp, mrp, created_by,
                      created_at, default_gst_rate, default_tax_mode
                    ) VALUES (
                      %s, %s, '', '', '', 'Imported Sales', 'Unit', 0, 0, 0, %s::jsonb, '[]'::jsonb,
                      'Imported from SALE INVOICE 28-MAR TO 28-APR-26.xlsx', '', '', %s, '', '', %s, %s, '', %s, '',
                      NULL, NULL, 'admin', %s, %s, %s
                    )
                    """,
                    [
                        name,
                        name,
                        json.dumps(sorted(product_usage[name])),
                        product_barcodes.get(name, ""),
                        name,
                        name,
                        name[:40],
                        min(row["date"] for row in rows if str(row["product_name"]) == name),
                        gst_rate,
                        tax_mode,
                    ],
                )
                products[name] = {
                    "sku": name,
                    "name": name,
                    "default_gst_rate": float(gst_rate),
                    "default_tax_mode": tax_mode,
                    "allowed_warehouse_ids_json": sorted(product_usage[name]),
                    "barcode": product_barcodes.get(name, ""),
                }
            else:
                existing_warehouses = set(products[name]["allowed_warehouse_ids_json"] or [])
                next_warehouses = sorted(existing_warehouses | product_usage[name])
                next_barcode = products[name]["barcode"] or product_barcodes.get(name, "")
                cur.execute(
                    "UPDATE products SET allowed_warehouse_ids_json = %s::jsonb, barcode = %s WHERE sku = %s",
                    [json.dumps(next_warehouses), next_barcode, name],
                )
                products[name]["allowed_warehouse_ids_json"] = next_warehouses
                products[name]["barcode"] = next_barcode

        for shop_name in sorted(shops):
            cur.execute(
                """
                INSERT INTO counterparties (
                  id, type, name, gst_number, bank_name, bank_account_number, ifsc_code, mobile_number,
                  address, city, delivery_address, delivery_city, contact_person, latitude, longitude,
                  location_label, created_by, created_at
                ) VALUES (
                  %s, 'Shop', %s, 'N/A', 'N/A', 'N/A', 'N/A', '', '', '', '', '', %s, NULL, NULL, NULL, 'admin', %s
                )
                """,
                [shop_id(shop_name), shop_name, shop_name, min(row["date"] for row in rows if str(row["shop_name"]) == shop_name)],
            )

        cur.execute(
            """
            SELECT warehouse_id, product_sku, COALESCE(SUM(quantity_available), 0) AS available_qty
            FROM inventory_lots
            GROUP BY warehouse_id, product_sku
            """
        )
        available_qty = {(str(row["warehouse_id"]), str(row["product_sku"])): as_float(row["available_qty"]) for row in cur.fetchall()}

        required_qty: dict[tuple[str, str], float] = defaultdict(float)
        for row in rows:
            key = (str(row["warehouse_id"]), str(row["product_name"]))
            required_qty[key] += as_float(row["qty"])

        earliest_sale = min(row["date"] for row in rows)
        opening_created_at = earliest_sale - timedelta(days=1)
        opening_lot_count = 0
        for (warehouse_id, product_sku), needed in sorted(required_qty.items()):
            have = available_qty.get((warehouse_id, product_sku), 0)
            short = round(needed - have, 6)
            if short <= 0:
                continue
            cur.execute(
                """
                INSERT INTO inventory_lots (
                  lot_id, source_order_id, source_type, warehouse_id, product_sku,
                  quantity_available, quantity_reserved, quantity_blocked, status, created_at
                ) VALUES (%s, %s, 'Opening', %s, %s, %s, 0, 0, 'Available', %s)
                """,
                [
                    make_opening_lot_id(warehouse_id, product_sku),
                    f"OPENING-STOCK-{warehouse_id}",
                    warehouse_id,
                    product_sku,
                    short,
                    opening_created_at,
                ],
            )
            opening_lot_count += 1

        sales_order_count = 0
        ledger_count = 0
        for invoice in sorted(invoice_groups, key=lambda key: invoice_meta[key]["date"]):
            group = invoice_groups[invoice]
            meta = invoice_meta[invoice]
            shop_name = str(meta["shop_name"])
            created_at = meta["date"]
            shop_key = shop_id(shop_name)
            invoice_total = 0.0
            for index, row in enumerate(group, start=1):
                product_name = str(row["product_name"])
                product = products[product_name]
                amount = as_float(row["amount"])
                qty = as_float(row["qty"])
                if qty <= 0:
                    continue
                is_gst = str(row["bill_type"]) == "GST"
                default_gst_rate = as_float(product.get("default_gst_rate"))
                gst_rate = default_gst_rate if is_gst and default_gst_rate > 0 else 0.0
                if is_gst and gst_rate > 0:
                    taxable_amount = round(amount / (1 + (gst_rate / 100)), 2)
                    gst_amount = round(amount - taxable_amount, 2)
                    tax_mode = "Inclusive"
                else:
                    taxable_amount = round(amount, 2)
                    gst_amount = 0.0
                    tax_mode = "NA"
                rate = round(taxable_amount / qty, 6)
                line_id = make_sales_line_id(invoice, index)
                note_parts = [
                    "Imported from SALE INVOICE 28-MAR TO 28-APR-26.xlsx",
                    f"Invoice {invoice}",
                    f"Warehouse source {row['warehouse_id']}",
                    f"Bill type {row['bill_type']}",
                    "Payment not imported. Collection will be tagged later.",
                ]
                cur.execute(
                    """
                    INSERT INTO sales_orders (
                      id, cart_id, shop_id, product_sku, salesman_id, warehouse_id, quantity, rate,
                      taxable_amount, gst_rate, gst_amount, tax_mode, total_amount, payment_mode,
                      cash_timing, delivery_mode, delivery_charge, note, status, created_at
                    ) VALUES (
                      %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Cash', NULL,
                      'Self Collection', 0, %s, 'Delivered', %s
                    )
                    """,
                    [
                        line_id,
                        invoice,
                        shop_key,
                        product_name,
                        sales_user["id"],
                        row["warehouse_id"],
                        qty,
                        rate,
                        taxable_amount,
                        gst_rate,
                        gst_amount,
                        tax_mode,
                        round(amount, 2),
                        " | ".join(note_parts),
                        created_at,
                    ],
                )
                consume_inventory(cur, str(row["warehouse_id"]), product_name, qty)
                sales_order_count += 1
                invoice_total += round(amount, 2)

            cur.execute(
                """
                INSERT INTO ledger_entries (
                  id, side, linked_order_id, party_name, goods_value, paid_amount, pending_amount, status, created_at
                ) VALUES (%s, 'Sales', %s, %s, %s, 0, %s, 'Pending', %s)
                """,
                [
                    f"IMP-LED-SALES-{re.sub(r'[^A-Za-z0-9]+', '-', invoice).strip('-')}",
                    invoice,
                    shop_name,
                    round(invoice_total, 2),
                    round(invoice_total, 2),
                    created_at,
                ],
            )
            ledger_count += 1

        conn.commit()
        print("Backup saved to", BACKUP_PATH)
        print("Imported shop counterparties:", len(shops))
        print("Imported sales invoices:", len(invoice_groups))
        print("Imported sales lines:", sales_order_count)
        print("Inserted opening lots:", opening_lot_count)
        print("Inserted sales ledger entries:", ledger_count)
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
